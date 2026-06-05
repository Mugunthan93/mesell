#!/usr/bin/env python3
"""
Meesho XLSX category template parser — v0.1

Parses XLSX files in data/meesho_templates/ to extract per-leaf field schemas.

For each file, extracts:
- leaf_id (from filename suffix)
- sheet inventory
- main upload sheet (`*-Fill this`) field schema:
    - name
    - marker ('compulsory' | 'recommended' | 'optional')
    - data_type ('text' | 'number' | 'dropdown' | 'image_url' | 'date')
    - help_text
    - enum_values (truncated to 50 if larger)
    - enum_count + enum_truncated flag
    - max_length (if textLength validation present)

Discovered structure (probed on Sarees / Mobile Covers / Spices — uniform):
- Sheet 1: 'Instructions'        — uniform across all templates
- Sheet 2: '{Cat}-Fill this'     — main upload sheet (variance lives here)
- Sheet 3: 'Example Sheet'       — sample rows
- Sheet 4: 'Validation Sheet'    — dropdown enum source
- Sheet 5: 'Return Reasons'      — uniform

Main sheet layout:
- Row 1: template title cell
- Row 2: per-column marker — '* Compulsory Field' | 'Recommended Field' | 'Optional Field' | 'Do not fill...'
- Row 3: '\\n\\n{field_name}\\n\\n{help_text}\\n' — split on newlines
- Row 4+: data rows (formulas)
- User-field columns start at column 4 (cols 1-3 are meta: Field Names label, ERROR STATUS, ERROR MESSAGE)

Usage:
    python3 scripts/parse_meesho_xlsx.py --super-ids 11 29 \\
        --output data/parsed/batch_01_women_fashion.json

    python3 scripts/parse_meesho_xlsx.py --leaves 10003 10382 14366 \\
        --output data/parsed/sample_validation.json
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = ROOT / "data" / "meesho_templates"
TREE_FILE = ROOT / "backend" / "app" / "data" / "meesho_category_tree.json"

FIELD_COL_START = 4
LARGE_ENUM_TRUNCATION = 50

COMPULSORY_PAT = re.compile(r"\*\s*compulsor", re.I)
RECOMMENDED_PAT = re.compile(r"recommend", re.I)
OPTIONAL_PAT = re.compile(r"^optional", re.I)
SKIP_PAT = re.compile(r"do not fill", re.I)

# v0.2: tightened to canonical image-field patterns only.
# Catches "Image 1", "Image 2 (Front)", "Front Image", "Back Image"
# Excludes "Still Image Sensor Resolution" and other tech-spec fields containing "Image".
IMAGE_FIELD_PAT = re.compile(
    r"^image\s+\d+(\s*\(.+\))?$|^(front|back|side|top|bottom|main|primary)\s+image$",
    re.I,
)

NUMBER_NAME_KEYWORDS = (
    "price", "qty", "quantity", "weight", "length", "width", "height",
    "size in", "mrp", "inventory", "voltage", "wattage", "frequency",
    "capacity", "diameter", "depth",
)


def find_main_sheet(wb):
    for n in wb.sheetnames:
        if n.endswith("-Fill this") or n.endswith("Fill this"):
            return n
    return None


def parse_marker(text):
    if text is None:
        return None
    s = str(text)
    if SKIP_PAT.search(s):
        return "meta"
    if COMPULSORY_PAT.search(s):
        return "compulsory"
    if RECOMMENDED_PAT.search(s):
        return "recommended"
    if OPTIONAL_PAT.search(s):
        return "optional"
    return None


def parse_field_description(text):
    """Row 3 cell: '\\n\\n{name}\\n\\n{help_text}\\n' (typical). Returns (name, help_text)."""
    if text is None:
        return None, None
    s = str(text)
    parts = [p.strip() for p in re.split(r"\n+", s) if p.strip()]
    if not parts:
        return None, None
    name = parts[0]
    help_text = "\n".join(parts[1:]) if len(parts) > 1 else None
    return name, help_text


def deref_validation(wb, formula1):
    """Resolve a list-validation formula → (values, source_label)."""
    if not formula1:
        return None, "no_formula"
    s = formula1.strip()
    if s.startswith("="):
        s = s[1:]
    if "!" not in s and "," in s:
        if s.startswith('"') and s.endswith('"'):
            s = s[1:-1]
        return [v.strip() for v in s.split(",") if v.strip()], "literal"
    m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", s)
    if not m:
        return None, f"unparsable:{s[:60]}"
    sheet_name, c1, r1, c2, r2 = m.groups()
    if sheet_name not in wb.sheetnames:
        return None, f"missing_sheet:{sheet_name}"
    vs = wb[sheet_name]
    c1_idx = column_index_from_string(c1)
    vals = []
    for r in range(int(r1), int(r2) + 1):
        v = vs.cell(row=r, column=c1_idx).value
        if v is None or v == "":
            continue
        vals.append(str(v).strip())
    return vals, "ranged"


def infer_data_type(field_name, enum_values, help_text):
    name = field_name or ""
    name_l = name.lower()
    # v0.2: canonical image patterns only (not loose substring match)
    if IMAGE_FIELD_PAT.match(name):
        return "image_url"
    if enum_values:
        return "dropdown"
    if any(k in name_l for k in NUMBER_NAME_KEYWORDS):
        return "number"
    if "date" in name_l or "expiry" in name_l:
        return "date"
    return "text"


def parse_xlsx(path: Path):
    anomalies = []
    record = {
        "filename": path.name,
        "leaf_id": None,
        "main_sheet": None,
        "field_count": 0,
        "compulsory_count": 0,
        "recommended_count": 0,
        "optional_count": 0,
        "fields": [],
        "sheets": [],
    }
    m = re.search(r"__(\d+)\.xlsx$", path.name)
    if not m:
        anomalies.append({"file": path.name, "reason": "no_leaf_id_in_filename"})
        return record, anomalies
    record["leaf_id"] = m.group(1)

    try:
        wb = openpyxl.load_workbook(str(path), data_only=False, read_only=False)
    except Exception as e:
        anomalies.append({"file": path.name, "reason": f"load_error:{e!s:.120}"})
        return record, anomalies

    record["sheets"] = [
        {"name": n, "max_row": wb[n].max_row, "max_col": wb[n].max_column} for n in wb.sheetnames
    ]

    main_name = find_main_sheet(wb)
    if not main_name:
        anomalies.append({"file": path.name, "reason": "no_main_sheet"})
        return record, anomalies
    record["main_sheet"] = main_name
    ws = wb[main_name]

    col_dvs = {}
    for dv in ws.data_validations.dataValidation:
        ranges = dv.sqref.ranges if hasattr(dv.sqref, "ranges") else [dv.sqref]
        for r in ranges:
            try:
                bounds = range_boundaries(str(r))
                if not bounds:
                    continue
                min_col, _, max_col, _ = bounds
                for c in range(min_col, max_col + 1):
                    col_dvs.setdefault(c, []).append(dv)
            except Exception:
                continue

    for c in range(FIELD_COL_START, ws.max_column + 1):
        marker_cell = ws.cell(row=2, column=c).value
        marker = parse_marker(marker_cell)
        if marker is None or marker == "meta":
            continue
        desc_cell = ws.cell(row=3, column=c).value
        field_name, help_text = parse_field_description(desc_cell)
        if not field_name:
            anomalies.append(
                {"file": path.name, "reason": f"no_field_name:col{c}", "marker": str(marker_cell)[:40]}
            )
            continue

        enum_values = None
        enum_count = None
        enum_source = None
        max_length = None
        for dv in col_dvs.get(c, []):
            if dv.type == "list":
                vals, source = deref_validation(wb, dv.formula1)
                if vals is not None:
                    enum_values = vals
                    enum_count = len(vals)
                    enum_source = source
                    break
            elif dv.type == "textLength":
                try:
                    max_length = int(dv.formula1)
                except (TypeError, ValueError):
                    pass

        data_type = infer_data_type(field_name, enum_values, help_text)

        field = {
            "col": c,
            "name": field_name,
            "marker": marker,
            "data_type": data_type,
            "help_text": help_text,
        }
        if enum_count is not None:
            field["enum_count"] = enum_count
            field["enum_source"] = enum_source
            field["enum_values"] = enum_values[:LARGE_ENUM_TRUNCATION]
            if enum_count > LARGE_ENUM_TRUNCATION:
                field["enum_truncated"] = True
        if max_length is not None and max_length > 0:
            field["max_length"] = max_length

        record["fields"].append(field)
        record["field_count"] += 1
        if marker == "compulsory":
            record["compulsory_count"] += 1
        elif marker == "recommended":
            record["recommended_count"] += 1
        elif marker == "optional":
            record["optional_count"] += 1

    return record, anomalies


def load_leaves_for_super_ids(super_ids):
    tree = json.loads(TREE_FILE.read_text())
    super_set = {str(s) for s in super_ids}
    return [c for c in tree["categories"] if c.get("is_leaf") and str(c.get("super_id")) in super_set]


def find_xlsx_for_leaf(leaf_id):
    matches = list(TEMPLATES_DIR.glob(f"*__{leaf_id}.xlsx"))
    return matches[0] if matches else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--super-ids", nargs="+", help="Super-category ids to parse (filters meesho_category_tree.json)")
    ap.add_argument("--leaves", nargs="+", help="Specific leaf ids to parse")
    ap.add_argument("--output", required=True, help="Output JSON path")
    ap.add_argument("--limit", type=int, default=None, help="Stop after N files (for testing)")
    args = ap.parse_args()

    if args.leaves:
        targets = []
        for lid in args.leaves:
            xlsx = find_xlsx_for_leaf(lid)
            if xlsx:
                targets.append((lid, xlsx))
            else:
                print(f"WARN: no XLSX for leaf {lid}", file=sys.stderr)
    elif args.super_ids:
        leaves = load_leaves_for_super_ids(args.super_ids)
        targets = []
        for lf in leaves:
            xlsx = find_xlsx_for_leaf(lf["leaf_id"])
            if xlsx:
                targets.append((lf["leaf_id"], xlsx))
            else:
                print(
                    f"WARN: no XLSX for leaf {lf['leaf_id']} ({lf['leaf_name']})",
                    file=sys.stderr,
                )
    else:
        ap.error("Specify --super-ids or --leaves")

    if args.limit:
        targets = targets[: args.limit]

    print(f"Parsing {len(targets)} XLSX files...", file=sys.stderr)
    records = []
    all_anomalies = []
    failed = []
    for i, (lid, xlsx_path) in enumerate(targets, 1):
        rec, anoms = parse_xlsx(xlsx_path)
        if not rec.get("fields") or rec.get("main_sheet") is None:
            failed.append({"leaf_id": lid, "file": xlsx_path.name, "reason": "no_fields_or_main_sheet"})
        records.append(rec)
        all_anomalies.extend(anoms)
        if i % 25 == 0:
            print(f"  ... {i}/{len(targets)}", file=sys.stderr)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output = {
        "parser_version": "0.2",
        "parsed_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "target_count": len(targets),
        "parsed_count": len(records) - len(failed),
        "failed_count": len(failed),
        "failures": failed,
        "anomalies": all_anomalies,
        "leaves": records,
    }
    output_path.write_text(json.dumps(output, indent=2))
    print(
        f"DONE: parsed={output['parsed_count']}/{output['target_count']} "
        f"failed={output['failed_count']} → {output_path}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
