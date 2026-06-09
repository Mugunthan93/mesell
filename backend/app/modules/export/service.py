"""``export`` service — public + worker-internal helpers per §14.C.

Per BACKEND_ARCHITECTURE.md §14.C (LOCKED 2026-06-05).

Public surface (called from :mod:`.router`)
-------------------------------------------
* :func:`initiate_export` — POST /products/{id}/export-xlsx (6-step
  request flow per §14.B.1).
* :func:`get_export` — GET /exports/{id} (4-step polling flow per
  §14.B.2).

Cross-module surface (OPTIONAL V1.5 elevation per §2.D)
-------------------------------------------------------
* :func:`summary` — latest export per product_id; available but NOT
  consumed in V1 (dashboard matrix kept at 8 ✓).

Worker-internal helpers (called from :mod:`.tasks` only)
--------------------------------------------------------
* :func:`_run_export_pipeline` — 9-step orchestrator.
* :func:`_resolve_schema`, :func:`_select_strategy`, :func:`_build_row`,
  :func:`_apply_strategy`, :func:`_translate_enums`,
  :func:`_reorder_columns`, :func:`_restore_aliases`,
  :func:`_write_xlsx`, :func:`_round_trip_validate`,
  :func:`_package_images_zip`.

Cross-module call sites (§16.B + §16.C Rule 7)
==============================================

* :func:`catalog.service.assert_product_ownership`
* :func:`catalog.service.get_product_for_export`
* :func:`customer.service.get_compliance_block`
* :func:`category.service.fetch_schema`
* :func:`category.service.get_field_enum`
* :func:`image.service.list_images`

The Celery enqueue uses ``celery_app.send_task("export.xlsx", ...)``
per §16.C Rule 7 — task name as STRING, not direct task import.

GCS path conventions (LOCKED per §14.I + D9)
============================================
* XLSX: ``meesell-exports/{user_id}/{export_id}/sheet.xlsx``
* ZIP:  ``meesell-exports/{user_id}/{export_id}/images.zip``
"""

from __future__ import annotations

import logging
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Literal
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession

from app.adapters import gcs as gcs_adapter
from app.modules.catalog import service as catalog_service
from app.modules.category import service as category_service
from app.modules.customer import service as customer_service
from app.modules.export import repository as export_repo
from app.modules.export.domain import (
    CollapsedComplianceStrategy,
    ComplianceStrategy,
    ExportStatusSummary,
    RoundTripResult,
    StandardComplianceStrategy,
    XlsxColumnSpec,
    XlsxRowSpec,
)
from app.modules.export.exceptions import (
    ComplianceStrategyError,
    ExportEnumValidationError,
    ExportNotFoundError,
    FrontImageMissingError,
    ProductNotReadyForExportError,
    RoundTripValidationError,
    XlsxBuildError,
)
from app.modules.export.schemas import (
    ExportInitiatedResponse,
    ExportResponse,
)
from app.modules.image import service as image_service
from app.shared.config import settings
from app.shared.database import make_worker_session
from app.shared.valkey import get_valkey_otp

logger = logging.getLogger(__name__)

# Valkey hint key for D2 pending-window format echo.
_FORMAT_HINT_TTL_SECONDS = 600


def _format_hint_key(export_id: UUID) -> str:
    return f"export:format:{export_id}"


async def _set_format_hint(
    export_id: UUID,
    format: Literal["xlsx_only", "xlsx_with_images"],  # noqa: A002 — API contract
) -> None:
    """Write the D2 format hint to Valkey DB 0 with 10-min TTL.

    Best-effort — failures logged at WARNING but do not raise.  The
    hint is purely cosmetic for the brief pending window of the API
    response.
    """
    try:
        client = get_valkey_otp()
        await client.set(_format_hint_key(export_id), format, ex=_FORMAT_HINT_TTL_SECONDS)
    except Exception as exc:  # noqa: BLE001 — best-effort cosmetic hint
        logger.warning(
            "export.format_hint: Valkey write failed (export=%s): %r",
            export_id,
            exc,
        )


async def _read_format_hint(
    export_id: UUID,
) -> Literal["xlsx_only", "xlsx_with_images"] | None:
    """Read the D2 format hint from Valkey DB 0.  Best-effort."""
    try:
        client = get_valkey_otp()
        raw = await client.get(_format_hint_key(export_id))
    except Exception as exc:  # noqa: BLE001 — best-effort cosmetic hint
        logger.warning(
            "export.format_hint: Valkey read failed (export=%s): %r",
            export_id,
            exc,
        )
        return None
    if raw is None:
        return None
    val = raw.decode("utf-8") if isinstance(raw, (bytes, bytearray)) else str(raw)
    if val in ("xlsx_only", "xlsx_with_images"):
        return val  # type: ignore[return-value]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Public surface — initiate_export + get_export (router-callable)
#
# Signatures accept ``db`` positionally OR by keyword to match the
# router's ``await export_service.initiate_export(user_id=..., db=db)``
# call convention.
# ─────────────────────────────────────────────────────────────────────────────
async def initiate_export(
    user_id: UUID,
    product_id: UUID,
    request: Any,
    db: AsyncSession,
) -> ExportInitiatedResponse:
    """POST /products/{id}/export-xlsx — 6-step flow per §14.B.1.

    Steps:

    1. ``catalog.service.assert_product_ownership`` — 404 if not owned.
    2. ``catalog.service.get_product_for_export`` — verify
       ``status='ready'``; else :class:`ProductNotReadyForExportError`.
    3. If ``format='xlsx_with_images'``: verify
       ``image.service.list_images`` shows at least 1 ready image with
       ``idx=1``; else :class:`FrontImageMissingError`.
    4. Repository insert + write Valkey format hint.
    5. Enqueue Celery task ``"export.xlsx"`` via
       ``celery_app.send_task`` (name-based per §16.C Rule 7).
    6. Return :class:`ExportInitiatedResponse`.
    """
    fmt: Literal["xlsx_only", "xlsx_with_images"] = getattr(
        request, "format", "xlsx_with_images"
    )

    # Step 1 — ownership gate.
    await catalog_service.assert_product_ownership(product_id, user_id, db=db)

    # Step 2 — product readiness.
    snapshot = await catalog_service.get_product_for_export(
        product_id, user_id, db=db
    )
    if snapshot.validation_summary.status != "ready":
        raise ProductNotReadyForExportError()

    # Step 3 — front-image gate (only for xlsx_with_images).
    if fmt == "xlsx_with_images":
        images_payload = await image_service.list_images(
            user_id=user_id, product_id=product_id, db=db
        )
        image_summaries = list(getattr(images_payload, "images", []) or [])
        has_front_ready = any(
            getattr(img, "idx", None) == 1
            and getattr(img, "status", None) == "ready"
            for img in image_summaries
        )
        if not has_front_ready:
            raise FrontImageMissingError()

    # Step 4 — repository insert + Valkey hint.
    initiated_at = datetime.now(timezone.utc)
    export = await export_repo.insert(
        db=db,
        user_id=user_id,
        product_id=product_id,
        format=fmt,
        initiated_at=initiated_at,
    )
    await _set_format_hint(export.id, fmt)

    # Step 5 — enqueue Celery task via the task's own .delay() —
    # name binding is preserved at the @shared_task decorator
    # (name="export.xlsx") per §14.E + §14-EXPORT-D8.  Same pattern as
    # §11 image.service which calls ``image_precheck_task.delay(...)``.
    # Avoids importing ``app.workers.celery_app`` at request time (the
    # celery_app singleton reads ``settings.CELERY_BROKER_URL`` which
    # is set via env at test runtime — see tests/conftest.py).
    from app.modules.export.tasks import export_xlsx_task

    task = export_xlsx_task.delay(str(export.id), str(user_id))

    # Step 6 — compose the 202 response.
    return ExportInitiatedResponse(
        export_id=export.id,
        status="pending",
        enqueued_task_id=str(task.id),
        initiated_at=export.initiated_at,
    )


async def get_export(
    user_id: UUID,
    export_id: UUID,
    db: AsyncSession,
) -> ExportResponse:
    """GET /exports/{id} — 4-step flow per §14.B.2."""
    pending_hint = await _read_format_hint(export_id)

    export = await export_repo.find_by_id(
        db=db,
        user_id=user_id,
        export_id=export_id,
        pending_format_hint=pending_hint,
    )
    if export is None:
        raise ExportNotFoundError()

    xlsx_signed_url: str | None = None
    zip_signed_url: str | None = None
    if export.status == "ready" and export.xlsx_gcs_path:
        xlsx_signed_url = await gcs_adapter.generate_signed_url(
            path=export.xlsx_gcs_path,
            ttl_seconds=settings.GCS_SIGNED_URL_TTL_SECONDS,
            method="GET",
        )
        if export.zip_gcs_path:
            zip_signed_url = await gcs_adapter.generate_signed_url(
                path=export.zip_gcs_path,
                ttl_seconds=settings.GCS_SIGNED_URL_TTL_SECONDS,
                method="GET",
            )

    return ExportResponse(
        export_id=export.id,
        product_id=export.product_id,
        status=export.status,
        format=export.format,
        xlsx_signed_url=xlsx_signed_url,
        zip_signed_url=zip_signed_url,
        error_message=export.error_message,
        error_code=export.error_code,
        initiated_at=export.initiated_at,
        completed_at=export.completed_at,
        round_trip_validated=export.round_trip_validated,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Cross-module surface — V1 not consumed
# ─────────────────────────────────────────────────────────────────────────────
async def summary(
    user_id: UUID,
    product_ids: list[UUID],
    db: AsyncSession,
) -> dict[UUID, ExportStatusSummary]:
    """Latest export per product — OPTIONAL V1.5 dashboard elevation."""
    return await export_repo.summarize_by_products(db, user_id, product_ids)


# ─────────────────────────────────────────────────────────────────────────────
# Worker-internal helpers (called only from tasks.py; module-private)
# ─────────────────────────────────────────────────────────────────────────────
async def _run_export_pipeline(
    export_id: UUID,
    user_id: UUID,
) -> None:
    """Orchestrate the 9-step Export Adapter pipeline per §14.E.

    Opens its own worker session via :func:`make_worker_session`.
    On any typed exception → repository ``update_status_failed`` with
    the exception's ``error_code`` per §14.H.
    """
    try:
        async with make_worker_session() as db:
            pending_hint = await _read_format_hint(export_id)
            export_row = await export_repo.find_by_id(
                db, user_id, export_id, pending_format_hint=pending_hint
            )
            if export_row is None:
                logger.warning(
                    "export.pipeline: export %s not found under user %s — skipping",
                    export_id,
                    user_id,
                )
                return

            snapshot = await catalog_service.get_product_for_export(
                export_row.product_id, user_id, db=db
            )
            schema = await _resolve_schema(snapshot.category_id, db=db)

            compliance_shape = str(schema.get("compliance_shape") or "")
            strategy = _select_strategy(compliance_shape)

            row = await _build_row(
                product_id=export_row.product_id,
                user_id=user_id,
                schema=schema,
                snapshot=snapshot,
                db=db,
            )
            row = _apply_strategy(row, strategy, schema)
            row = await _translate_enums(row, snapshot.category_id, db=db)
            row = _reorder_columns(row, schema)
            row = _restore_aliases(row, schema)

            xlsx_bytes = _write_xlsx(row)

            rtv = _round_trip_validate(xlsx_bytes, row)
            if not rtv.passed:
                raise RoundTripValidationError(
                    detail=(
                        rtv.diagnostic
                        or f"Round-trip mismatch on fields: {list(rtv.mismatches)}"
                    )
                )

            xlsx_path = f"meesell-exports/{user_id}/{export_id}/sheet.xlsx"
            await gcs_adapter.upload_bytes(
                path=xlsx_path,
                data=xlsx_bytes,
                content_type=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

            zip_path: str | None = None
            if export_row.format == "xlsx_with_images":
                zip_bytes = await _package_images_zip(
                    image_refs=tuple(snapshot.image_refs or ()),
                    user_id=user_id,
                    db=db,
                )
                if zip_bytes:
                    zip_path = f"meesell-exports/{user_id}/{export_id}/images.zip"
                    await gcs_adapter.upload_bytes(
                        path=zip_path,
                        data=zip_bytes,
                        content_type="application/zip",
                    )

            completed_at = datetime.now(timezone.utc)
            await export_repo.update_status_ready(
                db=db,
                user_id=user_id,
                export_id=export_id,
                xlsx_gcs_path=xlsx_path,
                zip_gcs_path=zip_path,
                completed_at=completed_at,
                round_trip_validated=True,
            )
            await db.commit()
    except (
        ProductNotReadyForExportError,
        FrontImageMissingError,
        ExportEnumValidationError,
        ComplianceStrategyError,
        XlsxBuildError,
        RoundTripValidationError,
    ) as exc:
        await _persist_failure(
            export_id=export_id,
            user_id=user_id,
            error_message=str(getattr(exc, "detail", None) or exc.__class__.__name__),
            error_code=str(
                getattr(exc, "error_code", None)
                or getattr(exc, "code", None)
                or "unknown"
            ),
        )
    except Exception as exc:  # noqa: BLE001 — unknown failure
        logger.exception(
            "export.pipeline: unexpected failure (export=%s user=%s): %r",
            export_id,
            user_id,
            exc,
        )
        await _persist_failure(
            export_id=export_id,
            user_id=user_id,
            error_message=f"Unexpected error: {exc!r}",
            error_code="unknown",
        )


async def _persist_failure(
    *,
    export_id: UUID,
    user_id: UUID,
    error_message: str,
    error_code: str,
) -> None:
    """Persist a failed export row via its OWN worker session."""
    try:
        async with make_worker_session() as db:
            await export_repo.update_status_failed(
                db=db,
                user_id=user_id,
                export_id=export_id,
                error_message=error_message,
                error_code=error_code,
                completed_at=datetime.now(timezone.utc),
            )
            await db.commit()
    except Exception as exc:  # noqa: BLE001 — informational
        logger.warning(
            "export.pipeline: update_status_failed write failed "
            "(export=%s user=%s code=%s): %r",
            export_id,
            user_id,
            error_code,
            exc,
        )


async def _resolve_schema(
    category_id: UUID,
    *,
    db: AsyncSession,
) -> dict:
    """Step 1 — fetch the §5A.B 7-key envelope via cached
    :func:`category.service.fetch_schema`.
    """
    return await category_service.fetch_schema(category_id, db=db)


def _select_strategy(compliance_shape: str) -> ComplianceStrategy:
    """Step 2 — dispatch on ``schema.compliance_shape`` per §5A.F."""
    if compliance_shape == "standard":
        return StandardComplianceStrategy()
    if compliance_shape == "collapsed":
        return CollapsedComplianceStrategy()
    raise ComplianceStrategyError(
        detail=(
            f"Unknown compliance_shape '{compliance_shape}' — "
            "expected 'standard' or 'collapsed'."
        )
    )


def _value_from_snapshot(canonical_name: str, snapshot: Any) -> Any:
    """Resolve a canonical field value from the export snapshot.

    Precedence: ``fields_jsonb`` > ``ai_suggestions[name].value`` >
    ``ai_suggestions[name]``.  Returns ``""`` for a missing value so
    XLSX cells render as blank per §14.K fixture 12.
    """
    fields = getattr(snapshot, "fields", {}) or {}
    if canonical_name in fields and fields[canonical_name] is not None:
        return fields[canonical_name]
    ai = getattr(snapshot, "ai_suggestions", {}) or {}
    suggestion = ai.get(canonical_name)
    if isinstance(suggestion, dict) and "value" in suggestion:
        return suggestion.get("value", "") if suggestion.get("value") is not None else ""
    if suggestion is not None:
        return suggestion
    return ""


async def _build_row(
    *,
    product_id: UUID,  # noqa: ARG001 — kept for §14.C signature parity
    user_id: UUID,
    schema: dict,
    snapshot: Any,
    db: AsyncSession,
) -> XlsxRowSpec:
    """Step 3 — assemble an :class:`XlsxRowSpec` in CANONICAL column
    ordering.

    Per D7, ``meesho_column_header`` is sourced directly from
    ``schema["fields"][i].meesho_column_header`` (the seed pipeline
    pre-embeds the typo-preserved headers per ``MVP_ARCH §3``).
    """
    compliance_block = await customer_service.get_compliance_block(user_id, db)

    main_sheet_label = str(schema.get("main_sheet_label") or "Sheet1")
    fields_list = schema.get("fields") or []

    columns: list[XlsxColumnSpec] = []
    for idx, field_spec in enumerate(fields_list):
        if not isinstance(field_spec, dict):
            continue
        canonical = str(field_spec.get("canonical_name", ""))
        if not canonical:
            continue
        header = str(field_spec.get("meesho_column_header") or canonical)
        value = _value_from_snapshot(canonical, snapshot)
        columns.append(
            XlsxColumnSpec(
                canonical_name=canonical,
                meesho_column_header=header,
                meesho_column_index=idx,
                value=value,
            )
        )

    return XlsxRowSpec(
        main_sheet_label=main_sheet_label,
        columns=tuple(columns),
        compliance_block=compliance_block,
    )


def _build_compliance_maps(
    schema: dict,
    canonical_names: tuple[str, ...],
) -> tuple[dict[str, str], dict[str, int]]:
    """Extract per-canonical header + index maps for the strategy."""
    fields_list = schema.get("fields") or []
    header_map: dict[str, str] = {}
    index_map: dict[str, int] = {}
    for idx, field_spec in enumerate(fields_list):
        if not isinstance(field_spec, dict):
            continue
        canonical = str(field_spec.get("canonical_name", ""))
        if canonical in canonical_names:
            header_map[canonical] = str(
                field_spec.get("meesho_column_header") or canonical
            )
            index_map[canonical] = idx
    return header_map, index_map


def _apply_strategy(
    row: XlsxRowSpec,
    strategy: ComplianceStrategy,
    schema: dict,
) -> XlsxRowSpec:
    """Step 4 — invoke ``strategy.apply`` and merge into ``row.columns``.

    On any strategy exception → :class:`ComplianceStrategyError`.
    """
    # Lazy imports of the canonical name tuples (kept private inside
    # ``domain``; this is the one consumer that needs them).
    from app.modules.export.domain import (  # noqa: PLC0415
        _COLLAPSED_CANONICAL_FIELDS,
        _STANDARD_CANONICAL_FIELDS,
    )

    if row.compliance_block is None:
        return row

    try:
        if isinstance(strategy, StandardComplianceStrategy):
            canonicals: tuple[str, ...] = _STANDARD_CANONICAL_FIELDS
        elif isinstance(strategy, CollapsedComplianceStrategy):
            canonicals = _COLLAPSED_CANONICAL_FIELDS
        else:
            raise ComplianceStrategyError(
                detail=(
                    "Unknown ComplianceStrategy subclass: "
                    f"{type(strategy).__name__}"
                )
            )

        header_map, index_map = _build_compliance_maps(schema, canonicals)
        compliance_columns = strategy.apply(
            row.compliance_block,
            column_header_map=header_map,
            column_index_map=index_map,
        )
    except ComplianceStrategyError:
        raise
    except Exception as exc:  # noqa: BLE001 — strategy.apply contract
        raise ComplianceStrategyError(
            detail=(
                f"Compliance strategy {type(strategy).__name__} raised: {exc!r}"
            )
        ) from exc

    new_canonicals = {col.canonical_name for col in compliance_columns}

    if isinstance(strategy, StandardComplianceStrategy):
        # Replace existing compliance columns in row.columns by canonical
        # name match.  We do NOT append compliance columns that the schema
        # did NOT declare — the schema is the authoritative column inventory
        # per §5A.B + §14.K fixture 1 (the saree schema has 3 fields, NOT
        # the 9 LM fields; the strategy must not pollute the row with
        # phantom LM columns).
        merged: list[XlsxColumnSpec] = []
        for col in row.columns:
            if col.canonical_name in new_canonicals:
                replacement = next(
                    (c for c in compliance_columns if c.canonical_name == col.canonical_name),
                    None,
                )
                if replacement is not None:
                    merged.append(replacement)
                    continue
            merged.append(col)
        return XlsxRowSpec(
            main_sheet_label=row.main_sheet_label,
            columns=tuple(merged),
            compliance_block=row.compliance_block,
        )

    # Collapsed — drop the 9 LM canonicals, append the 3 derived columns.
    filtered = tuple(
        col for col in row.columns if col.canonical_name not in _STANDARD_CANONICAL_FIELDS
    )
    return XlsxRowSpec(
        main_sheet_label=row.main_sheet_label,
        columns=filtered + tuple(compliance_columns),
        compliance_block=row.compliance_block,
    )


async def _translate_enums(
    row: XlsxRowSpec,
    category_id: UUID,
    *,
    db: AsyncSession,
) -> XlsxRowSpec:
    """Step 5 — canonical → meesho translation per field.

    LAYER 3 HALLUCINATION GUARDRAIL per ``MVP_ARCH §9.7``.  Unknown
    canonical → :class:`ExportEnumValidationError`.
    """
    from app.modules.category.exceptions import (
        CategoryNotFoundError,
        FieldEnumNotFoundError,
    )

    new_columns: list[XlsxColumnSpec] = []
    for col in row.columns:
        if col.value in (None, ""):
            new_columns.append(col)
            continue
        try:
            payload = await category_service.get_field_enum(
                category_id, col.canonical_name, db
            )
        except (FieldEnumNotFoundError, CategoryNotFoundError):
            new_columns.append(col)
            continue
        entries = payload.get("enum_entries") or []
        if not entries:
            new_columns.append(col)
            continue
        canonical_to_meesho: dict[str, str] = {}
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            ckey = str(entry.get("canonical", ""))
            mkey = str(entry.get("meesho", ckey))
            if ckey:
                canonical_to_meesho[ckey] = mkey
        col_str = str(col.value)
        if col_str not in canonical_to_meesho:
            raise ExportEnumValidationError(
                detail=(
                    f"Unknown canonical enum '{col_str}' for field "
                    f"'{col.canonical_name}' in category {category_id}."
                )
            )
        translated_value = canonical_to_meesho[col_str]
        new_columns.append(
            XlsxColumnSpec(
                canonical_name=col.canonical_name,
                meesho_column_header=col.meesho_column_header,
                meesho_column_index=col.meesho_column_index,
                value=translated_value,
            )
        )

    return XlsxRowSpec(
        main_sheet_label=row.main_sheet_label,
        columns=tuple(new_columns),
        compliance_block=row.compliance_block,
    )


def _reorder_columns(row: XlsxRowSpec, schema: dict) -> XlsxRowSpec:  # noqa: ARG001
    """Step 6 — sort ``row.columns`` by ``meesho_column_index``."""
    sorted_columns = tuple(sorted(row.columns, key=lambda c: c.meesho_column_index))
    return XlsxRowSpec(
        main_sheet_label=row.main_sheet_label,
        columns=sorted_columns,
        compliance_block=row.compliance_block,
    )


def _restore_aliases(row: XlsxRowSpec, schema: dict) -> XlsxRowSpec:  # noqa: ARG001
    """Step 7 — restore canonical_name → meesho_column_header.

    DECISION FLAG §14-EXPORT-D7 — RUNTIME NO-OP.  The seed pipeline
    pre-embeds typo-preserved headers in
    ``templates.schema_jsonb.fields[*].meesho_column_header`` per
    ``MVP_ARCH §3``, so by the time this step runs the typo restoration
    has already been done in :func:`_build_row`.  The
    ``field_aliases.for_xlsx_export=TRUE`` rows are consumed at SEED
    time only; runtime does NOT query that table.

    This step is retained as an explicit no-op so the §14.C 9-step
    contract is structurally honored.  When V2 adds per-marketplace
    aliasing diverging from the seed-embedded headers, this is the
    insertion point.
    """
    return row


def _write_xlsx(row: XlsxRowSpec) -> bytes:
    """Step 8 — openpyxl write.  Raises :class:`XlsxBuildError` on
    openpyxl failure.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise XlsxBuildError(
            detail="openpyxl is not installed — install it to enable export."
        ) from exc

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = str(row.main_sheet_label or "Sheet1")[:31].replace(":", "-")
        headers = [col.meesho_column_header for col in row.columns]
        values = [col.value for col in row.columns]
        ws.append(headers)
        ws.append(values)
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except XlsxBuildError:
        raise
    except Exception as exc:  # noqa: BLE001 — openpyxl edge cases
        raise XlsxBuildError(detail=f"openpyxl write failed: {exc!r}") from exc


def _round_trip_validate(
    xlsx_bytes: bytes,
    original: XlsxRowSpec,
) -> RoundTripResult:
    """Step 9 — re-parse the just-written XLSX and assert canonical
    equivalence with ``original``.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return RoundTripResult(
            passed=False,
            mismatches=(),
            diagnostic=f"openpyxl not installed: {exc!r}",
        )

    try:
        wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:  # noqa: BLE001 — corrupt XLSX
        return RoundTripResult(
            passed=False,
            mismatches=(),
            diagnostic=f"Re-parse failed: {exc!r}",
        )

    if not rows or len(rows) < 2:
        return RoundTripResult(
            passed=False,
            mismatches=tuple(col.canonical_name for col in original.columns),
            diagnostic="Re-parsed XLSX has fewer than 2 rows.",
        )

    parsed_headers = list(rows[0])
    parsed_values = list(rows[1])

    expected_headers = [col.meesho_column_header for col in original.columns]
    if list(parsed_headers) != expected_headers:
        return RoundTripResult(
            passed=False,
            mismatches=tuple(col.canonical_name for col in original.columns),
            diagnostic=(
                f"Header mismatch.  Expected {expected_headers!r}, "
                f"got {list(parsed_headers)!r}."
            ),
        )

    mismatches: list[str] = []
    for col, parsed_val in zip(original.columns, parsed_values, strict=False):
        expected = col.value
        expected_norm = "" if expected is None else expected
        parsed_norm = "" if parsed_val is None else parsed_val
        if str(expected_norm) != str(parsed_norm):
            mismatches.append(col.canonical_name)

    if mismatches:
        return RoundTripResult(
            passed=False,
            mismatches=tuple(mismatches),
            diagnostic=f"Value mismatch on fields: {mismatches!r}",
        )
    return RoundTripResult(passed=True, mismatches=(), diagnostic=None)


async def _package_images_zip(
    *,
    image_refs: tuple[str, ...],
    user_id: UUID,
    db: AsyncSession,  # noqa: ARG001 — signature parity with cross-module surface
) -> bytes:
    """Image ZIP packaging — gather raw bytes for each GCS path and
    pack via :class:`zipfile.ZipFile.writestr`.

    Returns ``b""`` when ``image_refs`` is empty.  Best-effort per
    image — a single failed download is logged and skipped (the ZIP
    ships the images that did download).
    """
    if not image_refs:
        return b""

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for ref in image_refs:
            try:
                image_bytes = await gcs_adapter.download_bytes(path=ref)
            except Exception as exc:  # noqa: BLE001 — per-image best-effort
                logger.warning(
                    "export.zip_pack: failed to download %s (user=%s): %r",
                    ref,
                    user_id,
                    exc,
                )
                continue
            archive_name = ref.rsplit("/", 1)[-1] or ref
            zf.writestr(archive_name, image_bytes)

    return buffer.getvalue()


__all__ = [
    "get_export",
    "initiate_export",
    "summary",
]
