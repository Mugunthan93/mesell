"""Meesho Bulk Catalog Upload — Hair Accessories Template scraper (PoC).

Logs into supplier.meesho.com using WebKit (headed), navigates to the bulk
upload page, drills into the Women Fashion -> Accessories -> Hair Accessories
category path, and captures the XLSX template download.

Single-category PoC. On any block (403/429/captcha) or login failure the
script raises immediately — no retries.
"""

from __future__ import annotations

import logging
import os
import random
import re
import sys
import time
from pathlib import Path

from dotenv import load_dotenv
from playwright.sync_api import (
    Download,
    Page,
    Playwright,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path("/Users/mugunthansrinivasan/Project/mesell")
CREDS_FILE = PROJECT_ROOT / ".meesho_creds.env"

OUTPUT_DIR = Path("/tmp/meesho_templates")
SCREENSHOT_DIR = OUTPUT_DIR / "screenshots"
OUTPUT_FILE = OUTPUT_DIR / "hair_accessories.xlsx"

LOGIN_URL = "https://supplier.meesho.com/panel/v3/new/root/login"
BULK_UPLOAD_URL = "https://supplier.meesho.com/panel/v3/new/products/upload/bulk"

# Meesho's cascading category picker has 4 columns. The task brief asks for
# Women Fashion -> Accessories -> Hair Accessories and the 4th column then shows
# the leaf categories under Hair Accessories. We pick the leaf "Hair Accessories"
# (the directly-named first leaf under the Hair Accessories parent group).
CATEGORY_PATH = ["Women Fashion", "Accessories", "Hair Accessories", "Hair Accessories"]

NAV_TIMEOUT_MS = 45_000
ACTION_TIMEOUT_MS = 20_000
DOWNLOAD_TIMEOUT_MS = 60_000

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("meesho-scraper")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def jitter(min_s: float = 2.0, max_s: float = 3.5) -> None:
    """Sleep a small jittered interval to look human."""
    delay = random.uniform(min_s, max_s)
    log.debug("Sleeping %.2fs", delay)
    time.sleep(delay)


def screenshot(page: Page, name: str) -> Path:
    """Save a full-page screenshot under SCREENSHOT_DIR."""
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    path = SCREENSHOT_DIR / f"{name}.png"
    try:
        page.screenshot(path=str(path), full_page=True)
        log.info("Screenshot saved: %s", path)
    except Exception as exc:  # noqa: BLE001 — best-effort capture
        log.warning("Could not save screenshot %s: %s", name, exc)
    return path


def detect_block(page: Page) -> None:
    """Raise immediately if a captcha / rate-limit / 403 page is detected."""
    url = page.url
    if any(token in url for token in ("captcha", "/error", "/blocked")):
        raise RuntimeError(f"Detected block URL: {url}")

    body_text = ""
    try:
        body_text = (page.locator("body").inner_text(timeout=2_000) or "").lower()
    except Exception:  # noqa: BLE001
        return

    block_markers = (
        "access denied",
        "too many requests",
        "rate limit",
        "captcha",
        "are you a human",
        "forbidden",
    )
    for marker in block_markers:
        if marker in body_text:
            raise RuntimeError(f"Block marker detected in page body: {marker!r}")


def load_creds() -> tuple[str, str]:
    if not CREDS_FILE.exists():
        raise FileNotFoundError(f"Credentials file missing: {CREDS_FILE}")
    load_dotenv(CREDS_FILE)
    user = os.environ.get("MEESHO_USERNAME", "").strip()
    pwd = os.environ.get("MEESHO_PASSWORD", "").strip()
    if not user or not pwd:
        raise RuntimeError("MEESHO_USERNAME / MEESHO_PASSWORD missing in creds file")
    return user, pwd


# ---------------------------------------------------------------------------
# Login
# ---------------------------------------------------------------------------


def perform_login(page: Page, username: str, password: str) -> None:
    log.info("Navigating to login URL: %s", LOGIN_URL)
    page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT_MS)
    jitter()
    detect_block(page)

    # Username field — match by visible placeholder/label text
    user_field = None
    candidates = [
        lambda: page.get_by_placeholder(re.compile(r"email.*mobile", re.I)),
        lambda: page.get_by_label(re.compile(r"email.*mobile", re.I)),
        lambda: page.get_by_role("textbox", name=re.compile(r"email|mobile", re.I)),
        lambda: page.locator("input[type='text']").first,
    ]
    for build in candidates:
        try:
            loc = build()
            loc.wait_for(state="visible", timeout=5_000)
            user_field = loc
            break
        except Exception:  # noqa: BLE001
            continue
    if user_field is None:
        screenshot(page, "login_no_user_field")
        raise RuntimeError("Could not locate username field on login page")

    log.info("Filling username")
    user_field.fill(username)
    jitter(1.0, 1.8)

    pwd_field = None
    pwd_candidates = [
        lambda: page.get_by_placeholder(re.compile(r"password", re.I)),
        lambda: page.get_by_label(re.compile(r"password", re.I)),
        lambda: page.locator("input[type='password']").first,
    ]
    for build in pwd_candidates:
        try:
            loc = build()
            loc.wait_for(state="visible", timeout=5_000)
            pwd_field = loc
            break
        except Exception:  # noqa: BLE001
            continue
    if pwd_field is None:
        screenshot(page, "login_no_pwd_field")
        raise RuntimeError("Could not locate password field on login page")

    log.info("Filling password")
    pwd_field.fill(password)
    jitter(1.0, 1.8)

    # Submit
    submit = None
    submit_candidates = [
        lambda: page.get_by_role("button", name=re.compile(r"log\s*in|sign\s*in", re.I)),
        lambda: page.locator("button[type='submit']").first,
    ]
    for build in submit_candidates:
        try:
            loc = build()
            loc.wait_for(state="visible", timeout=5_000)
            submit = loc
            break
        except Exception:  # noqa: BLE001
            continue
    if submit is None:
        screenshot(page, "login_no_submit")
        raise RuntimeError("Could not locate login submit button")

    log.info("Submitting login form")
    submit.click()

    # Wait for URL to change off the login page
    try:
        page.wait_for_url(
            lambda u: "login" not in u, timeout=NAV_TIMEOUT_MS
        )
    except PlaywrightTimeoutError:
        screenshot(page, "login_no_redirect")
        # Look for an error toast
        body_text = ""
        try:
            body_text = page.locator("body").inner_text(timeout=2_000)
        except Exception:  # noqa: BLE001
            pass
        raise RuntimeError(
            f"Login did not redirect away from /login. URL={page.url!r} "
            f"body-snippet={body_text[:300]!r}"
        )

    detect_block(page)
    page.wait_for_load_state("domcontentloaded", timeout=NAV_TIMEOUT_MS)
    log.info("Login OK — current URL: %s", page.url)
    jitter()
    screenshot(page, "01_post_login")


# ---------------------------------------------------------------------------
# Navigation to Bulk Upload
# ---------------------------------------------------------------------------


def navigate_to_bulk_upload(page: Page) -> None:
    """Navigate to the Bulk Catalog Upload screen.

    Strategy:
      1. Try the direct URL first (fast path).
      2. If we land back on the dashboard / home, click the visible
         "Add Catalogs in Bulk" button.
      3. Fallback: open the "Catalog Uploads" item in the left sidebar.
    """
    log.info("Trying direct bulk upload URL: %s", BULK_UPLOAD_URL)
    page.goto(BULK_UPLOAD_URL, wait_until="domcontentloaded", timeout=NAV_TIMEOUT_MS)
    jitter()
    detect_block(page)

    if "login" in page.url:
        screenshot(page, "bulk_redirected_to_login")
        raise RuntimeError(f"Bulk-upload URL redirected back to login: {page.url}")

    page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT_MS)
    screenshot(page, "02a_after_direct_url")
    log.info("Direct nav landed at: %s", page.url)

    # Decide if we are actually on the bulk-upload screen.
    # Heuristic: page text mentions "Select Category" / "category" cascade,
    # or we still see the welcome dashboard.
    body_text = ""
    try:
        body_text = page.locator("body").inner_text(timeout=3_000).lower()
    except Exception:  # noqa: BLE001
        pass

    on_bulk = ("select category" in body_text) or ("women fashion" in body_text)

    if not on_bulk:
        log.info("Not on bulk-upload yet; clicking 'Add Catalogs in Bulk' button")
        clicked = False
        click_candidates = [
            lambda: page.get_by_role("button", name=re.compile(r"add\s+catalogs?\s+in\s+bulk", re.I)),
            lambda: page.get_by_role("link", name=re.compile(r"add\s+catalogs?\s+in\s+bulk", re.I)),
            lambda: page.get_by_text(re.compile(r"add\s+catalogs?\s+in\s+bulk", re.I)),
        ]
        for build in click_candidates:
            try:
                loc = build()
                count = loc.count()
                for idx in range(count):
                    cand = loc.nth(idx)
                    try:
                        if cand.is_visible():
                            cand.scroll_into_view_if_needed(timeout=ACTION_TIMEOUT_MS)
                            cand.click(timeout=ACTION_TIMEOUT_MS)
                            clicked = True
                            break
                    except Exception:  # noqa: BLE001
                        continue
                if clicked:
                    break
            except Exception:  # noqa: BLE001
                continue

        if not clicked:
            # Sidebar fallback
            log.info("'Add Catalogs in Bulk' not clickable — trying sidebar 'Catalog Uploads'")
            for build in (
                lambda: page.get_by_role("link", name=re.compile(r"catalog\s+uploads?", re.I)),
                lambda: page.get_by_text(re.compile(r"catalog\s+uploads?", re.I)),
            ):
                try:
                    loc = build()
                    if loc.count() > 0:
                        loc.first.click(timeout=ACTION_TIMEOUT_MS)
                        clicked = True
                        break
                except Exception:  # noqa: BLE001
                    continue

        if not clicked:
            screenshot(page, "bulk_nav_failed")
            raise RuntimeError(
                "Could not navigate to bulk-upload UI from dashboard"
            )

        jitter(2.0, 3.5)
        try:
            page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT_MS)
        except PlaywrightTimeoutError:
            pass

    screenshot(page, "02_bulk_upload_landed")
    log.info("Bulk upload page loaded at %s", page.url)


# ---------------------------------------------------------------------------
# Category cascading selection
# ---------------------------------------------------------------------------


def click_category(page: Page, label: str, column_index: int) -> None:
    """Click a category cell with the given label.

    The page renders cascading columns. We try multiple locator strategies
    because the exact markup is unknown without inspection.
    """
    log.info("Selecting category [col %d]: %s", column_index, label)

    # Strategy stack — try most-semantic first.
    strategies = [
        # Exact role+name match (button or treeitem or listitem)
        lambda: page.get_by_role("button", name=re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)),
        lambda: page.get_by_role("treeitem", name=re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)),
        lambda: page.get_by_role("listitem", name=re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)),
        lambda: page.get_by_role("menuitem", name=re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)),
        # Exact text match
        lambda: page.get_by_text(re.compile(rf"^\s*{re.escape(label)}\s*$", re.I)),
        # Looser fallback
        lambda: page.locator(f"text=/^\\s*{re.escape(label)}\\s*$/i").first,
    ]

    last_err: Exception | None = None
    for build in strategies:
        try:
            loc = build()
            count = loc.count()
            if count == 0:
                continue
            # Walk every match, keep visible ones, then pick the RIGHTMOST
            # (largest x-coordinate). This matters when the same label appears
            # both as a parent and as a leaf in different cascading columns.
            visible: list[tuple[float, object]] = []
            for idx in range(count):
                cand = loc.nth(idx)
                try:
                    if not cand.is_visible():
                        continue
                    box = cand.bounding_box()
                    x = box["x"] if box else 0.0
                    visible.append((x, cand))
                except Exception:  # noqa: BLE001
                    continue
            if not visible:
                continue
            visible.sort(key=lambda pair: pair[0])
            target = visible[-1][1]  # rightmost
            target.scroll_into_view_if_needed(timeout=ACTION_TIMEOUT_MS)
            target.click(timeout=ACTION_TIMEOUT_MS)
            jitter()
            detect_block(page)
            return
        except Exception as exc:  # noqa: BLE001
            last_err = exc
            continue

    screenshot(page, f"category_fail_{column_index}_{label.replace(' ', '_')}")
    raise RuntimeError(
        f"Could not click category {label!r} (column {column_index}): {last_err}"
    )


def select_category_path(page: Page) -> None:
    for idx, label in enumerate(CATEGORY_PATH):
        click_category(page, label, idx)
        # Give the next column time to populate
        page.wait_for_load_state("networkidle", timeout=ACTION_TIMEOUT_MS)
        jitter(1.5, 2.5)
    screenshot(page, "03_post_category_select")
    log.info("Category path selected: %s", " -> ".join(CATEGORY_PATH))


# ---------------------------------------------------------------------------
# Template download
# ---------------------------------------------------------------------------


def download_template(page: Page) -> Download:
    log.info("Looking for the template download link/button")

    # Scroll the right-hand panel so the download link is in view
    try:
        page.mouse.wheel(0, 600)
        jitter(0.8, 1.4)
    except Exception:  # noqa: BLE001
        pass

    # Strategy: match anything that says "download" near "template", or a link/button
    # in the "Don't have ... template?" section.
    strategies = [
        lambda: page.get_by_role("button", name=re.compile(r"download.*template|get.*template", re.I)),
        lambda: page.get_by_role("link", name=re.compile(r"download.*template|get.*template", re.I)),
        lambda: page.get_by_text(re.compile(r"download\s+template|get\s+template", re.I)),
        lambda: page.get_by_text(re.compile(r"download", re.I)),
    ]

    target = None
    last_err: Exception | None = None
    for build in strategies:
        try:
            loc = build()
            count = loc.count()
            for idx in range(count):
                cand = loc.nth(idx)
                try:
                    if cand.is_visible():
                        target = cand
                        break
                except Exception:  # noqa: BLE001
                    continue
            if target is not None:
                break
        except Exception as exc:  # noqa: BLE001
            last_err = exc

    if target is None:
        screenshot(page, "download_no_button")
        raise RuntimeError(
            f"Could not find a 'Download template' control. Last error: {last_err}"
        )

    log.info("Clicking download control")
    with page.expect_download(timeout=DOWNLOAD_TIMEOUT_MS) as dl_info:
        try:
            target.scroll_into_view_if_needed(timeout=ACTION_TIMEOUT_MS)
        except Exception:  # noqa: BLE001
            pass
        target.click(timeout=ACTION_TIMEOUT_MS)
    download = dl_info.value

    log.info(
        "Download captured: suggested_filename=%s, url=%s",
        download.suggested_filename,
        download.url,
    )
    download.save_as(str(OUTPUT_FILE))
    log.info("Saved download to %s", OUTPUT_FILE)
    jitter()
    screenshot(page, "04_post_download")
    return download


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------


def verify_xlsx() -> dict:
    if not OUTPUT_FILE.exists():
        raise RuntimeError(f"Expected XLSX missing: {OUTPUT_FILE}")
    size = OUTPUT_FILE.stat().st_size
    if size == 0:
        raise RuntimeError(f"Downloaded XLSX is empty: {OUTPUT_FILE}")

    from openpyxl import load_workbook  # local import

    wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    info: dict = {
        "path": str(OUTPUT_FILE),
        "size_bytes": size,
        "sheets": [],
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = ws.max_column or 0
        header_rows: list[list] = []
        # First 3 rows as preview
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
            header_rows.append(list(row))
            if row_idx >= 2:
                break
        info["sheets"].append(
            {
                "name": sheet_name,
                "max_column": max_col,
                "header_preview_rows": header_rows,
            }
        )
    wb.close()
    return info


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def run(pw: Playwright) -> dict:
    user, pwd = load_creds()

    browser = pw.webkit.launch(headless=False)
    context = browser.new_context(
        locale="en-IN",
        timezone_id="Asia/Kolkata",
        viewport={"width": 1440, "height": 900},
    )
    context.set_default_timeout(ACTION_TIMEOUT_MS)
    context.set_default_navigation_timeout(NAV_TIMEOUT_MS)
    page = context.new_page()

    summary: dict = {
        "login_url_after": None,
        "category_path": CATEGORY_PATH,
        "download_url": None,
        "download_suggested_filename": None,
        "output_file": str(OUTPUT_FILE),
        "screenshots": [],
    }

    try:
        perform_login(page, user, pwd)
        summary["login_url_after"] = page.url

        navigate_to_bulk_upload(page)
        select_category_path(page)
        download = download_template(page)
        summary["download_url"] = download.url
        summary["download_suggested_filename"] = download.suggested_filename
    finally:
        # Collect screenshot paths
        if SCREENSHOT_DIR.exists():
            summary["screenshots"] = sorted(
                str(p) for p in SCREENSHOT_DIR.glob("*.png")
            )
        context.close()
        browser.close()

    return summary


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)

    log.info("==== Meesho Hair Accessories template PoC starting ====")
    with sync_playwright() as pw:
        summary = run(pw)

    log.info("==== Scrape complete — verifying XLSX ====")
    verification = verify_xlsx()

    print("\n========= SUMMARY =========")
    for k, v in summary.items():
        print(f"{k}: {v}")
    print("\n========= XLSX VERIFICATION =========")
    print(f"path        : {verification['path']}")
    print(f"size_bytes  : {verification['size_bytes']}")
    for sheet in verification["sheets"]:
        print(f"\n-- sheet: {sheet['name']} (max_column={sheet['max_column']}) --")
        for ridx, row in enumerate(sheet["header_preview_rows"], start=1):
            print(f"  row {ridx}: {row}")
    print("==========================")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        log.exception("Scraper failed: %s", exc)
        sys.exit(1)
