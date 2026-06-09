"""§14.K integration test 1 — full pipeline happy path.

Drives the 9-step Export Adapter end-to-end with the cross-module
service surfaces mocked.  The XLSX bytes round-trip through openpyxl
back to a canonical dict and are validated.

GCS is mocked (no live bucket); the upload paths are inspected for
the §14.I + D9 path scheme.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock
from uuid import uuid4

import pytest

from app.modules.customer.domain import ComplianceBlock


@pytest.mark.asyncio
async def test_full_pipeline_happy_path(monkeypatch):
    """End-to-end happy path — mocks all 6 cross-module surfaces +
    GCS adapter; runs ``_run_export_pipeline``; verifies:

    * ``exports`` row reaches ``status='ready'``.
    * XLSX uploaded under the LOCKED path
      ``meesell-exports/{user_id}/{export_id}/sheet.xlsx``.
    * XLSX bytes re-parse via openpyxl and the header row matches the
      schema-supplied Meesho headers.
    """
    from app.modules.export import service as svc
    from app.modules.export.domain import Export
    from app.shared import database as shared_db

    user_id = uuid4()
    product_id = uuid4()
    category_id = uuid4()
    export_id = uuid4()

    fake_export = Export(
        id=export_id,
        user_id=user_id,
        product_id=product_id,
        format="xlsx_only",  # no ZIP path; happy-path baseline.
        status="pending",
        xlsx_gcs_path=None,
        zip_gcs_path=None,
        error_message=None,
        error_code=None,
        round_trip_validated=None,
        initiated_at=datetime.now(timezone.utc),
        completed_at=None,
    )

    # ── Mock the worker DB session ──────────────────────────────────────
    session_mock = AsyncMock()

    class _Ctx:
        async def __aenter__(self_inner):
            return session_mock

        async def __aexit__(self_inner, *args):
            return False

    monkeypatch.setattr(shared_db, "make_worker_session", lambda: _Ctx())

    # ── Mock cross-module surfaces ──────────────────────────────────────
    snapshot = SimpleNamespace(
        product_id=product_id,
        category_id=category_id,
        fields={"name": "Test Product"},
        ai_suggestions={},
        image_refs=(),
        validation_summary=SimpleNamespace(status="ready"),
    )

    compliance = ComplianceBlock(
        manufacturer_name="ACME",
        manufacturer_address="Addr",
        manufacturer_pincode="560001",
        packer_name="Pkr",
        packer_address="PA",
        packer_pincode="560002",
        importer_name=None,
        importer_address=None,
        importer_pincode=None,
        country_of_origin="IN",
    )

    schema = {
        "fields": [
            {
                "canonical_name": "name",
                "meesho_column_header": "Product Name",
                "data_type": "text",
                "primitive": "text_short",
                "marker": "compulsory",
                "enum_resolver": None,
                "is_advanced": False,
                "help_text": "h",
                "validation_message_ids": [],
            },
            {
                "canonical_name": "manufacturer_name",
                "meesho_column_header": "Manufacturer Name",
                "data_type": "text",
                "primitive": "text_short",
                "marker": "compulsory",
                "enum_resolver": None,
                "is_advanced": False,
                "help_text": "h",
                "validation_message_ids": [],
            },
        ],
        "compulsory_count": 2,
        "optional_count": 0,
        "total_count": 2,
        "wizard_step_count": 1,
        "main_sheet_label": "Happy Path Sheet",
        "compliance_shape": "standard",
    }

    # find_by_id returns the pending export.
    async def fake_find_by_id(db, uid, eid, *, pending_format_hint=None):
        return fake_export

    update_ready_calls = {"n": 0, "args": None}

    async def fake_update_status_ready(**kwargs):
        update_ready_calls["n"] += 1
        update_ready_calls["args"] = kwargs
        return fake_export

    async def fake_update_status_failed(**kwargs):
        raise AssertionError("update_status_failed should NOT be called on happy path")

    async def fake_get_product_for_export(pid, uid, db=None):
        return snapshot

    async def fake_get_compliance_block(uid, db):
        return compliance

    async def fake_fetch_schema(cid, db):
        return schema

    from app.modules.category.exceptions import FieldEnumNotFoundError

    async def fake_get_field_enum(cid, fname, db):
        # Treat both fields as non-enum (free-text) — pass-through.
        raise FieldEnumNotFoundError()

    monkeypatch.setattr(svc.export_repo, "find_by_id", fake_find_by_id)
    monkeypatch.setattr(svc.export_repo, "update_status_ready", fake_update_status_ready)
    monkeypatch.setattr(svc.export_repo, "update_status_failed", fake_update_status_failed)
    monkeypatch.setattr(svc.catalog_service, "get_product_for_export", fake_get_product_for_export)
    monkeypatch.setattr(svc.customer_service, "get_compliance_block", fake_get_compliance_block)
    monkeypatch.setattr(svc.category_service, "fetch_schema", fake_fetch_schema)
    monkeypatch.setattr(svc.category_service, "get_field_enum", fake_get_field_enum)

    # ── Mock GCS ───────────────────────────────────────────────────────
    gcs_uploads: dict[str, bytes] = {}

    async def fake_upload(path, data, content_type=None, **kw):
        gcs_uploads[path] = data

    monkeypatch.setattr(svc.gcs_adapter, "upload_bytes", fake_upload)

    # ── Bypass the Valkey hint reads (no live broker). ──────────────────
    async def fake_read_hint(eid):
        return "xlsx_only"

    monkeypatch.setattr(svc, "_read_format_hint", fake_read_hint)

    # ── Run the pipeline ───────────────────────────────────────────────
    await svc._run_export_pipeline(export_id, user_id)

    # ── Assertions ─────────────────────────────────────────────────────
    assert update_ready_calls["n"] == 1, "update_status_ready must fire exactly once"

    # Verify the GCS path scheme per §14.I + D9.
    expected_xlsx_path = f"meesell-exports/{user_id}/{export_id}/sheet.xlsx"
    assert expected_xlsx_path in gcs_uploads, (
        f"XLSX uploaded to wrong path. Got: {list(gcs_uploads.keys())}"
    )
    xlsx_bytes = gcs_uploads[expected_xlsx_path]
    assert len(xlsx_bytes) > 0

    # Re-parse the XLSX and verify the header row.
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) >= 2
    headers = list(rows[0])
    values = list(rows[1])
    assert headers == ["Product Name", "Manufacturer Name"]
    assert values == ["Test Product", "ACME"]
